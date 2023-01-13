#coding:windows-1251

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


db_type = dict[str, list[dict[str, str]]]
ml_type = Optional[dict[str, dict[str, str | list[dict[str, str | int | float]]]]]

now = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
markers = ['fmi']

target = Path('target')
ml_dir = Path('ml')
destin = ml_dir / now

for folder in target, ml_dir, destin:
    if not folder.exists():
        folder.mkdir()


def remove_cid(string: str) -> str:
    """Remove cid from string. """
    return re.sub(r'\(cid:\d+\)', '', string)


def cleaner(string: Optional[str]) -> Optional[str]:
    """Remove chars from string except particular set. 
    Return None if there's nothing to clean. 
    """
    if string:
        pattern = r'(?i)[^0-9a-zа-я.,_~-]'
        return re.sub(pattern, '', remove_cid(string))


def normalize_qty(qty: str) -> int | float:
    """Bring quantity to a common format. """
    if qty.isdigit():
        return int(qty)
    try:
        qty = float(qty.replace(',', '.'))
        return int(qty) if not qty % 1 else qty
    except ValueError:
        return 0  # as a marker of fail


def normalize_uom(uom: str) -> str:
    """Bring unit of measure to a common format. """
    # the only expected uoms are 'шт' and 'м', 
    # excessive pattern for unexpected 
    uom = re.sub(r'[^a-zа-я]', '', uom.lower())
    if uom == 'm': return 'м'  # if eng M for rus М 
    return uom


def collect_data(row: list[Optional[str]]) -> dict[str, str | int | float]:
    """Parse row and collect subitem data. """
    row = list(filter(bool, map(cleaner, row)))
    elem = row[-1].upper()
    qty = sum(map(normalize_qty, row[:2]))
    uom = ('шт' if len(row) == 3 
           else normalize_uom(row[2]))
    keys = 'elem', 'qty', 'uom'
    return dict(zip(keys, (elem, qty, uom)))


def get_validator(wb: Workbook, entries: list[str]) -> DataValidation:
    """Create a data validation object for dropdown list of given entries. """
    
    if any(entry.__contains__(',') for entry in entries):
        if 'dvranges' not in wb:
            wb.create_sheet('dvranges')
            wb['dvranges'].sheet_state = 'hidden'
        
        dvranges = wb['dvranges']
        column = 1 if not dvranges['A1'].value else dvranges.max_column + 1
        entries_number = len(entries)
        
        for row in range(1, entries_number + 1):
            dvranges.cell(row=row, column=column, value=entries[row - 1])
            
        letter = get_column_letter(column)
        formula = f"'dvranges'!${letter}$1:${letter}${entries_number}"
    else:
        formula = '"' + ','.join(entries) + '"'
    
    return DataValidation(type='list', formula1=formula, showErrorMessage=False)
