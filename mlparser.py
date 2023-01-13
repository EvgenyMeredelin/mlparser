#coding:windows-1251

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.worksheet import Worksheet

import tools as t


def parse_materials_lists(target: Path = t.target) -> t.ml_type:
    """Parse pdf files located in target folder and collect data to a dict. 
    Return None if folder is empty.
    """
    
    # if target folder is not empty
    if next(target.iterdir(), None):
        ml = {}  # create collected data container
        cache = set()  # create 'no duplicate kit' tester
        
        # filter pdf files and extract tables
        for file in target.glob('*.pdf'):
            fname = file.stem  # get filename without extension
            kit = re.split(r'[.\s]', fname)[0].upper()
            
            if kit not in cache and all(word not in fname for word in t.markers):
                cache.add(kit)  # to ignore duplicate kits titled differently
                with pdfplumber.open(file) as pdf:
                    tables = pdf.pages[0].extract_tables()
                
                # move parsed pdf to destination folder and note internal link
                path = file.rename(t.destin / file.name).name
                link = f'=HYPERLINK("{path}", "{kit}")'
                ml[kit] = {'link': link, 'subitems': []}
                
                # collect data from tables
                for table in tables:
                    for row in table:
                        if row[1] and row[0] and re.fullmatch(r'\d+\.?', row[0]):
                            ml[kit]['subitems'].append(t.collect_data(row[1:6]))
        
        # move remaining files from target to destination folder
        for file in target.glob('*'):
            file.rename(t.destin / file.name)
        
        return ml
    
    
def read_update_database() -> t.db_type:
    """Read, create/update a database in db.json. """
    
    # read existing or create a new database
    path = Path('db.json')
    
    if path.exists():
        with path.open(encoding='windows-1251') as file:
            db = json.load(file)
    else: 
        db = {}
    
    # update db with/create from unmarked valid records 
    # in excel loader file
    wb = openpyxl.load_workbook('db.xlsx')    
    
    for date, *row in wb.active:
        if not date.value:
            values = [cell.value for cell in row]
            
            if any(val is None for val in values):
                date.value = 'ValueError'
                continue
            
            code, name, uom = values
            keys = 'name', 'uom'
            record = dict(zip(keys, (name, uom)))
            
            if record not in db.setdefault(code, []):
                db[code].append(record)
            
            # mark processed record with a datetime
            date.value = t.now
            
    wb.save('db.xlsx')
    
    for code, records in db.items():
        db[code] = sorted(records, key=lambda r: r['name'])
    
    db = dict(sorted(db.items()))
    
    with path.open('w', encoding='windows-1251') as target:
        json.dump(db, target, ensure_ascii=False, indent=4)
    
    return db
    
    
def write_report(ml: t.ml_type, db: t.db_type) -> None:
    """Merge collected data with database records and write reports. """
    
    # SECTION 1: worksheets, constants, parameters, and helper objects
    # headers of materials lists report (MLR)
    MHEAD = ['Артикул ЗАКАЗЧИК\n', 'Наименование ПОСТАВЩИК\n',
             'Ед.\nизм.', 'Кол-\nво', 'Потребление\nв мес.',
             'Цена без НДС\nза комплект', 'Комментарий\n']
    
    # headers of er404 report
    EHEAD = ['Артикул ЗАКАЗЧИК\n', 'Входит\nв комплект(-ы)',
             'Вероятно,\nэто артикул', 'Наименование ПОСТАВЩИК\n',
             'Ед.\nизм.']
    
    # entries of manager notes dropdown list
    NOTES = ['внести изменения в базу', 'есть в базе', 'завести в базу',
             'завести в базу/заказать', 'запросить/есть в базе',
             'запросить/завести в базу', 'запросить/заказать']
    
    # entries of packets dropdown list
    PACKS = ['31 - Пакет 100x150', '35 - Пакет 200x250']
    
    IMPEL = 'Выберите значение из списка'  # select prompt
    ER404 = 'Не найдено'  # Not Found message
    codes = defaultdict(list)  # not found codes container
    LABEL = ['', 'Термоэтикетка  80Х30', 'шт', 1]  # kit label
    
    font_name, font_size = 'Consolas', 11 # common font params
    font = Font(name=font_name, size=font_size)  # default font
    link_font = Font(name=font_name, size=font_size, 
                     color='0000FF')  # hyperlink font
    head_font = Font(name=font_name, size=font_size,
                     italic=True)  # reports head row font
    
    HEIGHT_UNIT = font_size + 3  # basic row height
    WIDTH_UNIT = 8/5 * HEIGHT_UNIT  # basic column width
    
    side = Side(style='thin', color='626567')  # cell side
    border = Border(top=side, bottom=side, 
                    left=side, right=side)  # cell border
    
    # create workbook with MLR worksheet
    wb = openpyxl.Workbook()
    mlrep = wb.active
    mlrep.title = 'Матлисты'
    
    # make a local copy of database in a hidden worksheet
    database = wb.create_sheet('database')
    database.sheet_state = 'hidden'
    for code, records in db.items():
        for record in records:
            database.append([code, *record.values()])
    
    # create multiple choice worksheet
    wsmul = wb.create_sheet('Мультивыбор')
    wsmul.append(MHEAD[:2])
    cache = set()  # 'no duplicate code' tester
    
    # some inner helper functions
    get_row = lambda ws: ws.max_row + 1
    
    def get_name_uom(code: str, ws: Worksheet, column: int) -> tuple[str, str]:
        """Get name/uom or dropdown list of names/uom vlookup formula pair. """
        
        # chinese items granted highest priority
        china = [record for record in db[code]
                 if record['name'].__contains__('CHN')]
        
        # if no chinese items suggest all maintaining the code
        records = china or db[code]
        
        if len(records) == 1:
            name, uom = records[0].values()
        else:
            names = [record['name'] for record in records]
            dv = t.get_validator(wb, names)
            
            if code not in cache:
                wsmul.add_data_validation(dv)
                name_mul = Cell(wsmul, row=get_row(wsmul),
                                column=2, value=IMPEL)
                dv.add(name_mul)
                wsmul.append([code, name_mul])
                cache.add(code)  # to ignore further code occurrences
            
            ws.add_data_validation(dv)
            code_col_letter = t.get_column_letter(column - 1)
            curr_row = get_row(ws)
            value = (f"=VLOOKUP({code_col_letter}{curr_row}, "
                     + "'Мультивыбор'!A:B, 2, 0)")
            name = Cell(ws, row=get_row(ws), column=column,
                        value=value)
            dv.add(name)
            name_col_letter = t.get_column_letter(column)
            uom = (f'=IFERROR(VLOOKUP({name_col_letter}{curr_row}, '
                   + 'database!B:C, 2, 0), "")')
        return name, uom
    
    # SECTION 2: make and format materials lists report
    mlrep.append(MHEAD)
    
    # add packets dropdown list to MLR worksheet
    dv_packs = t.get_validator(wb, PACKS)
    mlrep.add_data_validation(dv_packs)
    
    for kit, data in ml.items():
        link, subitems = data.values()
        mlrep.append([link, f'Комплект крепежа {kit}', 'упак', 1])
        
        for subitem in subitems:
            code, qty, uom = subitem.values()
            
            if code in db:
                mlrep.append([code, *get_name_uom(code, mlrep, 2), qty])
            else:
                # if code not in database get uom from pdf source file
                mlrep.append([code, ER404, uom, qty])
                codes[code].append(kit)
        
        # finalize kit with packets dropdown list, label, and blank row
        packet = Cell(mlrep, row=get_row(mlrep), column=2, value=PACKS[-1])
        dv_packs.add(packet)
        packet = ['', packet, 'шт', 1]
        for row in packet, LABEL, []:
            mlrep.append(row)
            
    # add manager notes dropdown list to MLR worksheet
    dv_notes = t.get_validator(wb, NOTES)
    dv_notes.add(f'G2:G{mlrep.max_row}')  # comments column
    mlrep.add_data_validation(dv_notes)
    
    # add conditional formatting to MLR worksheet
    # format kit head row
    style_grn = DifferentialStyle(font=Font(bold=True, color='145A32'),
                                  fill=PatternFill(bgColor='A9DFBF'))
    ruleKhead = Rule(type='expression', dxf=style_grn, stopIfTrue=True)
    ruleKhead.formula = ['$C2="упак"']
    
    # format er404 message
    style_red = DifferentialStyle(font=Font(italic=True, color='9C0006'),
                                  fill=PatternFill(bgColor='FFC7CE'))
    ruleEr404 = Rule(type='containsText', operator='containsText',
                     text=ER404, dxf=style_red)
    
    # format select prompt
    style_yel = DifferentialStyle(font=Font(italic=True, color='7D6608'),
                                  fill=PatternFill(bgColor='F9E79F'))
    ruleImpel = Rule(type='containsText', operator='containsText',
                     text=IMPEL, dxf=style_yel)
    
    for rule in ruleKhead, ruleEr404, ruleImpel:
        mlrep.conditional_formatting.add(f'A2:G{mlrep.max_row}', rule)
    
    # SECTION 3: make and format er404 report
    ws404 = wb.create_sheet(title=ER404)
    ws404.append(EHEAD)
    
    for code, kits in sorted(codes.items()):
        guess = ('0' + code if code.isdigit()  # case: first 0 truncated
                 else 'FO' + code[2:] if code[:2] == 'F0'  # case: misprint 
                 else '')  # for 'guess not in database' branch
        
        subrow = ([guess, *get_name_uom(guess, ws404, 4)] if guess in db 
                  else [code, '', ''])  # code, name, uom
        
        flag = True  # first or the only kit for the current code
        
        for kit in kits:
            link = Cell(ws404, value=ml[kit]['link'])
            row = [code, link] + subrow if flag else ['', link]
            ws404.append(row)
            flag = False
            
    # format select prompt in multiple choice worksheet and er404 report
    for ws, col in [(wsmul, 'B'), (ws404, 'D')]:
        ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}', ruleImpel)
    
    # SECTION 4: apply cross-worksheet formatting
    # fit rows and columns (adjust dimensions)
    column_width_resolver = {
        mlrep: [('A', 1), ('B', 3.25), ('C', .25), ('D', .25), ('E', .75),
                ('F', .75), ('G', 1.25)],
        wsmul: [('A', 1), ('B', 3.25)],
        ws404: [('A', 1), ('B', 1), ('C', 1), ('D', 3.25), ('E', .25)]
    }
    
    for sheet, column_width in column_width_resolver.items():
        for letter, coeff in column_width:
            sheet.column_dimensions[letter].width = coeff * WIDTH_UNIT
        
        sheet.row_dimensions[1].height = 2 * HEIGHT_UNIT
        for row in range(2, get_row(sheet)):
            sheet.row_dimensions[row].height = HEIGHT_UNIT
            
        # apply fonts and borders
        rows = tuple(sheet.rows)
        for cell in rows[0]:
            cell.border = border
            cell.font = head_font
        for row in rows[1:]:
            for cell in row:
                cell.border = border
                islink = str(cell.value).startswith('=HYPERLINK')
                cell.font = link_font if islink else font
        
        # freeze first row (above A2)
        sheet.freeze_panes = 'A2'
                
    wb.save(t.destin / f'{t.now}.xlsx')


if __name__ == '__main__':
    ml = parse_materials_lists()
    db = read_update_database()
    
    if ml and db:
        write_report(ml, db)
    